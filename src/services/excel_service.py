"""Excel service for import/export operations"""

from typing import List, Optional, Dict, Any
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from ..models.subject import Subject
from ..models.lesson import Lesson
from ..utils.logger import setup_logger

logger = setup_logger()


class ExcelService:
    """Service for Excel import/export operations"""
    
    @staticmethod
    def import_subject_from_excel(file_path: str) -> Optional[Subject]:
        """Import subject from Excel file"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Read subject basic info (assuming first few rows)
            subject_name = sheet.cell(row=1, column=2).value or ""
            subject_code = sheet.cell(row=2, column=2).value
            location = sheet.cell(row=3, column=2).value
            default_duration = sheet.cell(row=4, column=2).value
            category_main = sheet.cell(row=5, column=2).value
            category_sub = sheet.cell(row=6, column=2).value
            
            if not subject_name:
                logger.error("Subject name is required")
                return None
            
            # Read lessons (assuming starting from row 8)
            lessons = []
            row = 8
            while True:
                lesson_name = sheet.cell(row=row, column=1).value
                if not lesson_name:
                    break
                
                lesson_duration = sheet.cell(row=row, column=2).value
                materials = sheet.cell(row=row, column=3).value
                
                lesson = Lesson(
                    name=str(lesson_name),
                    duration=float(lesson_duration) if lesson_duration else None
                )
                
                if materials:
                    # Materials might be comma-separated file paths
                    lesson.materials = [m.strip() for m in str(materials).split(",")]
                
                lessons.append(lesson)
                row += 1
            
            # Create subject
            subject = Subject(
                name=subject_name,
                code=str(subject_code) if subject_code else None,
                location=str(location) if location else None,
                default_duration=float(default_duration) if default_duration else None,
                category_main=str(category_main) if category_main else None,
                category_sub=str(category_sub) if category_sub else None,
                lessons=lessons
            )
            
            logger.info(f"Successfully imported subject '{subject_name}' from Excel")
            return subject
            
        except Exception as e:
            logger.error(f"Error importing subject from Excel: {e}")
            return None
    
    @staticmethod
    def create_subject_template(file_path: str):
        """Create Excel template for subject import"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Subject Template"
            
            # Headers
            sheet.cell(row=1, column=1, value="Tên môn học:")
            sheet.cell(row=2, column=1, value="Mã môn học:")
            sheet.cell(row=3, column=1, value="Địa điểm:")
            sheet.cell(row=4, column=1, value="Thời lượng mặc định (giờ):")
            sheet.cell(row=5, column=1, value="Phân loại chính:")
            sheet.cell(row=6, column=1, value="Phân loại phụ:")
            
            # Lesson headers
            sheet.cell(row=7, column=1, value="Tên bài học")
            sheet.cell(row=7, column=2, value="Thời lượng (giờ)")
            sheet.cell(row=7, column=3, value="Tài liệu (đường dẫn, phân cách bằng dấu phẩy)")
            
            # Save
            workbook.save(file_path)
            logger.info(f"Created subject template at {file_path}")
            
        except Exception as e:
            logger.error(f"Error creating subject template: {e}")
    
    @staticmethod
    def export_schedule_to_excel(schedule, file_path: str) -> bool:
        """Export schedule to Excel file"""
        try:
            from ..models.schedule import Schedule
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Thời khóa biểu"
            
            # Header
            sheet.cell(row=1, column=1, value="Thời khóa biểu")
            if schedule.name:
                sheet.cell(row=1, column=2, value=schedule.name)
            
            row = 3
            
            # Write each week
            for week in schedule.weeks:
                # Week header
                sheet.cell(row=row, column=1, value=f"Tuần {week.week_number}")
                sheet.cell(row=row, column=2, value=f"{week.start_date} - {week.end_date}")
                row += 1
                
                # Day headers
                sheet.cell(row=row, column=1, value="Ngày")
                sheet.cell(row=row, column=2, value="Thời gian")
                sheet.cell(row=row, column=3, value="Môn học")
                sheet.cell(row=row, column=4, value="Bài học")
                sheet.cell(row=row, column=5, value="Địa điểm")
                row += 1
                
                # Write each day
                for day in week.days:
                    if not day.items:
                        continue
                    
                    for item in day.items:
                        sheet.cell(row=row, column=1, value=day.date.strftime("%Y-%m-%d"))
                        sheet.cell(row=row, column=2, value=f"{item.start_time} - {item.end_time}")
                        sheet.cell(row=row, column=3, value=item.subject_name)
                        sheet.cell(row=row, column=4, value=item.lesson_name)
                        sheet.cell(row=row, column=5, value=item.location or "")
                        row += 1
                    
                    row += 1  # Empty row between days
            
            workbook.save(file_path)
            logger.info(f"Successfully exported schedule to Excel: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting schedule to Excel: {e}")
            return False

